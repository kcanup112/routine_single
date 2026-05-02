"""
Email routes using SMTP2GO API — sends routine as Excel attachment
"""
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
from typing import List, Optional
import httpx
import logging
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.core.database_saas import get_db
from app.core.config_saas import settings
from app.models.models import (
    Teacher, User, ClassRoutineEntry, Day, Period, Department, Subject, Class
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/email", tags=["email"])

SMTP2GO_API_URL = "https://api.smtp2go.com/v3/email/send"
SENDER_EMAIL = "anupkc@kec.edu.np"


def _send_via_smtp2go(to_email: str, subject: str, html_body: str, attachments: list = None):
    """Send a single email via SMTP2GO API. Raises on failure."""
    payload = {
        "sender": SENDER_EMAIL,
        "to": [to_email],
        "subject": subject,
        "html_body": html_body,
    }
    if attachments:
        payload["attachments"] = attachments

    resp = httpx.post(
        SMTP2GO_API_URL,
        json=payload,
        headers={
            "Content-Type": "application/json",
            "X-Smtp2go-Api-Key": settings.MAIL_API_KEY,
        },
        timeout=30,
    )
    data = resp.json()
    if resp.status_code != 200 or data.get("data", {}).get("failed", 0) > 0:
        error_msg = data.get("data", {}).get("error", resp.text)
        raise Exception(error_msg)
    return data


class BulkEmailRequest(BaseModel):
    teacher_ids: Optional[List[int]] = None  # None means all teachers


class UserEmailRequest(BaseModel):
    subject: str
    html: str
    user_ids: Optional[List[int]] = None


# ── helpers ────────────────────────────────────────────────────────────

def _build_teacher_routine_map(entries, teacher_id, all_teachers):
    """Build a {day_id-period_id: data} map for one teacher."""
    routine_map = {}
    total_load = 0.0

    for e in entries:
        is_lead = e.lead_teacher_id == teacher_id
        is_a1 = e.assist_teacher_1_id == teacher_id
        is_a2 = e.assist_teacher_2_id == teacher_id
        is_a3 = e.assist_teacher_3_id == teacher_id

        if not (is_lead or is_a1 or is_a2 or is_a3):
            continue

        periods = e.num_periods or 1
        load = periods * (0.8 if e.is_lab else 1.0)
        total_load += load

        partners = []
        if e.is_lab:
            ids = [e.lead_teacher_id, e.assist_teacher_1_id,
                   e.assist_teacher_2_id, e.assist_teacher_3_id]
            partners = [tid for tid in ids if tid and tid != teacher_id]

        partner_names = []
        for pid in partners:
            t = next((t for t in all_teachers if t.id == pid), None)
            if t:
                partner_names.append(t.abbreviation or t.name)

        subj = e.subject
        cls = e.class_

        key = f"{e.day_id}-{e.period_id}"
        routine_map[key] = {
            "subject": subj.name if subj else "N/A",
            "code": subj.code if subj else "",
            "class_name": cls.name if cls else "N/A",
            "section": cls.section if cls else "",
            "room_no": cls.room_no if cls else "",
            "is_lab": e.is_lab,
            "num_periods": periods,
            "role": "Lead" if is_lead else "Assistant",
            "partners": partner_names,
        }

    return routine_map, total_load


def _generate_routine_excel(teacher, department_name, routine_map, days, periods, total_load):
    """Return bytes of an xlsx workbook for one teacher."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Routine"

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 — college name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periods) + 1)
    c = ws.cell(row=1, column=1, value="Kantipur Engineering College")
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(periods) + 1)
    c = ws.cell(row=2, column=1, value="Teacher's Routine")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3 — department + effective info
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    ws.cell(row=3, column=1, value=f"Department: {department_name}").font = Font(bold=True)

    # Row 4 — teacher name + total load
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    ws.cell(row=4, column=1, value=f"Teacher: {teacher.name}").font = Font(bold=True)
    if len(periods) >= 8:
        ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=min(11, len(periods) + 1))
        ws.cell(row=4, column=9, value=f"Total Load: {total_load:.1f}").font = Font(bold=True)

    # Row 5 — spacer
    ws.row_dimensions[5].height = 6

    # Row 6 — header
    header_fill = PatternFill("solid", fgColor="D3D3D3")
    hdr = ws.cell(row=6, column=1, value="Days \\ Time")
    hdr.font = Font(bold=True)
    hdr.fill = header_fill
    hdr.border = border_all
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, period in enumerate(periods):
        start = period.start_time.strftime("%H:%M") if period.start_time else ""
        end = period.end_time.strftime("%H:%M") if period.end_time else ""
        c = ws.cell(row=6, column=idx + 2, value=f"{start}-{end}")
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = border_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    day_fill = PatternFill("solid", fgColor="F0F0F0")
    for d_idx, day in enumerate(days):
        row = 7 + d_idx
        ws.row_dimensions[row].height = 72

        dc = ws.cell(row=row, column=1, value=day.name)
        dc.font = Font(bold=True)
        dc.fill = day_fill
        dc.border = border_all
        dc.alignment = Alignment(horizontal="center", vertical="center")

        skip = 0
        for p_idx, period in enumerate(periods):
            col = p_idx + 2
            if skip > 0:
                skip -= 1
                continue

            key = f"{day.id}-{period.id}"
            data = routine_map.get(key)

            if not data:
                c = ws.cell(row=row, column=col, value="-")
                c.border = border_all
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                text = data["subject"]
                if data["is_lab"]:
                    text += " (Lab)"
                text += f"\n[{data['class_name']}"
                if data["section"]:
                    text += f" - {data['section']}"
                text += "]"
                if data["room_no"]:
                    text += f"\nRoom: {data['room_no']}"
                if data["partners"]:
                    text += f"\nWith: {', '.join(data['partners'])}"

                c = ws.cell(row=row, column=col, value=text)
                c.border = border_all
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if data["num_periods"] > 1:
                    end_col = col + data["num_periods"] - 1
                    if end_col <= len(periods) + 1:
                        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
                        for mc in range(col, end_col + 1):
                            ws.cell(row=row, column=mc).border = border_all
                    skip = data["num_periods"] - 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── routes ─────────────────────────────────────────────────────────────

@router.post("/send-routines-to-teachers")
def send_routines_to_teachers(payload: BulkEmailRequest, db: Session = Depends(get_db)):
    """Generate each teacher's routine as Excel and email it as attachment."""
    if not settings.MAIL_API_KEY:
        raise HTTPException(status_code=500, detail="MAIL_API_KEY (SMTP2GO) not configured")

    # Fetch reference data
    days = db.query(Day).order_by(Day.day_number).all()
    periods = db.query(Period).filter(Period.is_teaching_period == True).order_by(Period.period_number).all()
    all_teachers = db.query(Teacher).filter(Teacher.is_active == True).all()
    departments = {d.id: d.name for d in db.query(Department).all()}

    # Fetch all routine entries (with relationships)
    entries = db.query(ClassRoutineEntry).options(
        joinedload(ClassRoutineEntry.subject),
        joinedload(ClassRoutineEntry.class_),
    ).all()

    # Teachers to email
    query = db.query(Teacher).filter(
        Teacher.email.isnot(None),
        Teacher.email != "",
        Teacher.is_active == True,
    )
    if payload.teacher_ids:
        query = query.filter(Teacher.id.in_(payload.teacher_ids))
    teachers_to_mail = query.all()

    if not teachers_to_mail:
        raise HTTPException(status_code=404, detail="No teachers with email found")

    email_body = (
        "<p>Please find your routine attached with this email.</p>"
        "<br/>"
        "<p>Sincerely,<br/>"
        "Department of Computer and Electronics Engineering<br/>"
        "Kantipur Engineering College</p>"
    )

    sent = []
    failed = []
    for teacher in teachers_to_mail:
        try:
            dept_name = departments.get(teacher.department_id, "Not Assigned")
            routine_map, total_load = _build_teacher_routine_map(entries, teacher.id, all_teachers)
            xlsx_bytes = _generate_routine_excel(teacher, dept_name, routine_map, days, periods, total_load)
            xlsx_b64 = base64.b64encode(xlsx_bytes).decode("utf-8")

            filename = f"Routine_{teacher.name.replace(' ', '_')}.xlsx"

            _send_via_smtp2go(
                to_email=teacher.email,
                subject="Your Teaching Routine - Kantipur Engineering College",
                html_body=email_body,
                attachments=[{
                    "filename": filename,
                    "fileblob": xlsx_b64,
                    "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }],
            )
            sent.append({"id": teacher.id, "name": teacher.name, "email": teacher.email})
        except Exception as e:
            logger.error(f"Failed to send routine to {teacher.email}: {e}")
            failed.append({"id": teacher.id, "name": teacher.name, "email": teacher.email, "error": str(e)})

    return {"sent": len(sent), "failed": len(failed), "sent_list": sent, "failed_list": failed}


@router.post("/send-to-users")
def send_email_to_users(payload: UserEmailRequest, db: Session = Depends(get_db)):
    """Send email to users (all or selected by IDs)"""
    if not settings.MAIL_API_KEY:
        raise HTTPException(status_code=500, detail="MAIL_API_KEY (SMTP2GO) not configured")

    query = db.query(User).filter(
        User.email.isnot(None),
        User.email != "",
        User.is_active == True,
    )
    if payload.user_ids:
        query = query.filter(User.id.in_(payload.user_ids))

    users = query.all()

    if not users:
        raise HTTPException(status_code=404, detail="No users with email found")

    sent = []
    failed = []
    for user in users:
        try:
            _send_via_smtp2go(
                to_email=user.email,
                subject=payload.subject,
                html_body=payload.html,
            )
            sent.append({"id": user.id, "name": user.full_name, "email": user.email})
        except Exception as e:
            logger.error(f"Failed to send email to {user.email}: {e}")
            failed.append({"id": user.id, "name": user.full_name, "email": user.email, "error": str(e)})

    return {"sent": len(sent), "failed": len(failed), "sent_list": sent, "failed_list": failed}

    return {"sent": len(sent), "failed": len(failed), "sent_list": sent, "failed_list": failed}
